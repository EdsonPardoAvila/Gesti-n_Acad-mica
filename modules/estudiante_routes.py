from flask import request, jsonify, send_file
import pandas as pd
from datetime import datetime
import os
import sqlite3
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from .database import get_db_connection

UPLOAD_FOLDER = 'uploads'
REPORT_FOLDER = 'reportes'
ALLOWED_EXTENSIONS = {'csv', 'xlsx', 'xls'}

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(REPORT_FOLDER, exist_ok=True)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def init_estudiante_routes(app):
    
    @app.route('/api/estudiantes/cargar', methods=['POST'])
    def cargar_estudiantes():
        try:
            if 'file' not in request.files:
                return jsonify({'success': False, 'message': 'No se encontró el archivo'}), 400
            
            file = request.files['file']
            if file.filename == '':
                return jsonify({'success': False, 'message': 'Nombre de archivo vacío'}), 400
            if not allowed_file(file.filename):
                return jsonify({'success': False, 'message': 'Formato no permitido'}), 400
            
            filename = secure_filename(file.filename)
            filepath = os.path.join(UPLOAD_FOLDER, filename)
            file.save(filepath)
            
            # Leer el archivo
            if filename.endswith('.csv'):
                df = pd.read_csv(filepath, encoding='utf-8-sig')
            else:
                df = pd.read_excel(filepath)
            
            # Mostrar columnas originales para depuración
            print("Columnas originales del archivo:", df.columns.tolist())
            
            # Normalizar nombres de columnas
            df.columns = [col.strip().lower().replace(' ', '_').replace('.', '').replace('(', '').replace(')', '') for col in df.columns]
            
            print("Columnas normalizadas:", df.columns.tolist())
            
            # Mapeo de columnas más amplio
            mapeo_columnas = {
                'nombre_estudiante': ['nombre_estudiante', 'nombreestudiante', 'nombre', 'nombres', 'estudiante', 'nombre_del_estudiante'],
                'numero_documento': ['n_documento', 'numero_documento', 'documento', 'identificacion', 'cedula', 'id_estudiante', 'num_documento', 'documento_identidad'],
                'programa': ['programa', 'carrera', 'curso', 'facultad', 'programa_academico'],
                'correo_personal': ['correo_personal', 'correopersonal', 'email_personal', 'email', 'correo_electronico'],
                'correo_institucional': ['correo_institucional', 'correoinstitucional', 'email_institucional', 'email_ institucional'],
                'telefono': ['telefono', 'celular', 'telefono_movil', 'movil', 'whatsapp', 'teléfono', 'cel', 'numero_telefono', 'phone', 'contacto'],
                'dependencia': ['dependencia', 'area', 'facultad', 'departamento', 'dependencia_encargada'],
                'accion': ['accion', 'proceso', 'gestion', 'actividad', 'acción'],
                'estado': ['estado', 'status', 'situacion', 'estado_proceso'],
                'evidencia': ['evidencia', 'prueba', 'soporte', 'documento', 'evidencia_proceso'],
                'asistencia': ['asistencia', 'estado_asistencia', 'presente', 'asistio'],
                'observaciones': ['observaciones', 'obs', 'notas', 'comentarios', 'observacion']
            }
            
            renombrar = {}
            for col_destino, posibles_nombres in mapeo_columnas.items():
                for nombre in posibles_nombres:
                    if nombre in df.columns:
                        renombrar[nombre] = col_destino
                        print(f"Mapeado: '{nombre}' -> '{col_destino}'")
                        break
            
            df.rename(columns=renombrar, inplace=True)
            
            print("Columnas después del mapeo:", df.columns.tolist())
            
            # Verificar si existe la columna telefono
            if 'telefono' not in df.columns:
                print("ADVERTENCIA: No se encontró columna de teléfono. Creando columna vacía.")
                df['telefono'] = ''
            
            conn = get_db_connection()
            registros_insertados = 0
            registros_actualizados = 0
            
            for index, row in df.iterrows():
                cursor = conn.cursor()
                documento = str(row.get('numero_documento', ''))
                if documento and documento != 'nan':
                    # Procesar teléfono - asegurar que sea string y limpiar
                    telefono_raw = row.get('telefono', '')
                    if pd.isna(telefono_raw):
                        telefono = ''
                    else:
                        telefono = str(telefono_raw).strip()
                        # Limpiar caracteres no numéricos pero mantener el formato
                        # Solo eliminar espacios y caracteres especiales, mantener números y signo +
                        if telefono and telefono != 'nan':
                            # Si comienza con +, mantenerlo (código internacional)
                            if telefono.startswith('+'):
                                telefono = '+' + ''.join(filter(lambda x: x.isdigit(), telefono[1:]))
                            else:
                                telefono = ''.join(filter(str.isdigit, telefono))
                    
                    print(f"Fila {index}: Documento={documento}, Teléfono original={row.get('telefono')}, Teléfono procesado={telefono}")
                    
                    cursor.execute('SELECT id FROM estudiantes WHERE numero_documento = ?', (documento,))
                    existe = cursor.fetchone()
                    
                    if existe:
                        cursor.execute('''
                            UPDATE estudiantes SET 
                                nombre_estudiante=?, 
                                programa=?, 
                                correo_personal=?,
                                correo_institucional=?, 
                                telefono=?,
                                dependencia=?,
                                accion=?,
                                estado=?,
                                evidencia=?,
                                asistencia=?, 
                                observaciones=?
                            WHERE numero_documento=?
                        ''', (
                            str(row.get('nombre_estudiante', '')), 
                            str(row.get('programa', '')),
                            str(row.get('correo_personal', '')), 
                            str(row.get('correo_institucional', '')),
                            telefono,
                            str(row.get('dependencia', '')),
                            str(row.get('accion', '')),
                            str(row.get('estado', 'Pendiente')),
                            str(row.get('evidencia', '')),
                            str(row.get('asistencia', '')),
                            str(row.get('observaciones', '')), 
                            documento
                        ))
                        registros_actualizados += 1
                    else:
                        cursor.execute('''
                            INSERT INTO estudiantes (
                                nombre_estudiante, 
                                numero_documento, 
                                programa,
                                correo_personal, 
                                correo_institucional, 
                                telefono,
                                dependencia,
                                accion,
                                estado,
                                evidencia,
                                asistencia, 
                                observaciones
                            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        ''', (
                            str(row.get('nombre_estudiante', '')), 
                            documento,
                            str(row.get('programa', '')), 
                            str(row.get('correo_personal', '')),
                            str(row.get('correo_institucional', '')), 
                            telefono,
                            str(row.get('dependencia', '')),
                            str(row.get('accion', '')),
                            str(row.get('estado', 'Pendiente')),
                            str(row.get('evidencia', '')),
                            str(row.get('asistencia', '')), 
                            str(row.get('observaciones', ''))
                        ))
                        registros_insertados += 1
            
            conn.commit()
            conn.close()
            os.remove(filepath)
            
            return jsonify({
                'success': True, 
                'message': f'Carga exitosa. {registros_insertados} nuevos, {registros_actualizados} actualizados.'
            })
        except Exception as e:
            print(f"Error en carga masiva: {str(e)}")
            import traceback
            traceback.print_exc()
            return jsonify({'success': False, 'message': str(e)}), 500
    
    @app.route('/api/estudiantes', methods=['GET'])
    def obtener_estudiantes():
        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM estudiantes ORDER BY fecha_creacion DESC, id DESC')
            estudiantes = [dict(row) for row in cursor.fetchall()]
            conn.close()
            return jsonify({'success': True, 'data': estudiantes, 'total': len(estudiantes)})
        except Exception as e:
            return jsonify({'success': False, 'message': str(e)}), 500

    @app.route('/api/estudiantes/<int:id>', methods=['PUT'])
    def actualizar_estudiante(id):
        try:
            data = request.json
            conn = get_db_connection()
            cursor = conn.cursor()
            
            telefono = data.get('telefono', '')
            if telefono:
                telefono = ''.join(filter(str.isdigit, telefono))
            
            cursor.execute('''
                UPDATE estudiantes 
                SET nombre_estudiante=?, 
                    programa=?, 
                    correo_personal=?,
                    correo_institucional=?, 
                    telefono=?,
                    dependencia=?,
                    accion=?,
                    estado=?,
                    evidencia=?,
                    asistencia=?, 
                    observaciones=?
                WHERE id=?
            ''', (
                data.get('nombre_estudiante', ''), 
                data.get('programa', ''),
                data.get('correo_personal', ''), 
                data.get('correo_institucional', ''),
                telefono,
                data.get('dependencia', ''),
                data.get('accion', ''),
                data.get('estado', 'Pendiente'),
                data.get('evidencia', ''),
                data.get('asistencia', ''),
                data.get('observaciones', ''), 
                id
            ))
            conn.commit()
            conn.close()
            return jsonify({'success': True, 'message': 'Estudiante actualizado correctamente'})
        except Exception as e:
            return jsonify({'success': False, 'message': str(e)}), 500

    @app.route('/api/estudiantes/eliminar/<int:id>', methods=['DELETE'])
    def eliminar_estudiante(id):
        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute('DELETE FROM estudiantes WHERE id = ?', (id,))
            conn.commit()
            conn.close()
            return jsonify({'success': True, 'message': 'Registro eliminado exitosamente'})
        except Exception as e:
            return jsonify({'success': False, 'message': str(e)}), 500

    @app.route('/api/estudiantes/actualizar-estado/<int:id>', methods=['PATCH'])
    def actualizar_estado_estudiante(id):
        try:
            data = request.json
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute('''
                UPDATE estudiantes 
                SET estado=?, accion=?, evidencia=?
                WHERE id=?
            ''', (
                data.get('estado', 'Pendiente'),
                data.get('accion', ''),
                data.get('evidencia', ''),
                id
            ))
            conn.commit()
            conn.close()
            return jsonify({'success': True, 'message': 'Estado actualizado correctamente'})
        except Exception as e:
            return jsonify({'success': False, 'message': str(e)}), 500

    @app.route('/api/estudiantes/reporte', methods=['GET'])
    def descargar_reporte_estudiantes():
        try:
            conn = get_db_connection()
            df = pd.read_sql_query('SELECT * FROM estudiantes ORDER BY fecha_creacion DESC, id DESC', conn)
            conn.close()
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Reporte Estudiantes"
            
            header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
            header_fill = PatternFill(start_color='198754', end_color='198754', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                top=Side(style='thin'), bottom=Side(style='thin'))
            
            columns = ['ID', 'Nombre Estudiante', 'N. Documento', 'Programa', 'Correo Personal',
                       'Correo Institucional', 'Teléfono', 'Dependencia', 'Acción', 'Estado', 
                       'Evidencia', 'Asistencia', 'Observaciones', 'Fecha Creación']
            
            for col, header in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            for row_idx, row in df.iterrows():
                ws.cell(row=row_idx + 2, column=1, value=row['id'])
                ws.cell(row=row_idx + 2, column=2, value=row['nombre_estudiante'] or '')
                ws.cell(row=row_idx + 2, column=3, value=row['numero_documento'] or '')
                ws.cell(row=row_idx + 2, column=4, value=row['programa'] or '')
                ws.cell(row=row_idx + 2, column=5, value=row['correo_personal'] or '')
                ws.cell(row=row_idx + 2, column=6, value=row['correo_institucional'] or '')
                ws.cell(row=row_idx + 2, column=7, value=row['telefono'] or '')
                ws.cell(row=row_idx + 2, column=8, value=row['dependencia'] or '')
                ws.cell(row=row_idx + 2, column=9, value=row['accion'] or '')
                ws.cell(row=row_idx + 2, column=10, value=row['estado'] or 'Pendiente')
                ws.cell(row=row_idx + 2, column=11, value=row['evidencia'] or '')
                ws.cell(row=row_idx + 2, column=12, value=row['asistencia'] or '')
                ws.cell(row=row_idx + 2, column=13, value=row['observaciones'] or '')
                ws.cell(row=row_idx + 2, column=14, value=row['fecha_creacion'] or '')
            
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            filename = f'reporte_estudiantes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            filepath = os.path.join(REPORT_FOLDER, filename)
            wb.save(filepath)
            
            return send_file(filepath, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                            as_attachment=True, download_name=filename)
        except Exception as e:
            return jsonify({'success': False, 'message': str(e)}), 500

    @app.route('/api/estudiantes/plantilla', methods=['GET'])
    def descargar_plantilla_estudiantes():
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Plantilla Estudiantes"
            
            header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
            header_fill = PatternFill(start_color='198754', end_color='198754', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            columns = ['Nombre Estudiante', 'N. Documento', 'Programa', 'Correo Personal',
                       'Correo Institucional', 'Teléfono', 'Dependencia', 'Acción', 
                       'Estado', 'Evidencia', 'Asistencia', 'Observaciones']
            
            for col, header in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            ejemplo = ['María González', '123456789', 'Ingeniería de Sistemas', 
                      'maria@gmail.com', 'maria@universidad.edu', '3001234567',
                      'Bienestar Universitario', 'Asesoría Académica', 'En Proceso',
                      'Soporte_adjunto.pdf', 'Presente', 'Estudiante regular']
            
            for col, value in enumerate(ejemplo, 1):
                ws.cell(row=2, column=col, value=value)
            
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            filename = 'plantilla_estudiantes.xlsx'
            filepath = os.path.join(REPORT_FOLDER, filename)
            wb.save(filepath)
            
            return send_file(filepath, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                            as_attachment=True, download_name='Plantilla_Carga_Estudiantes.xlsx')
        except Exception as e:
            return jsonify({'success': False, 'message': str(e)}), 500
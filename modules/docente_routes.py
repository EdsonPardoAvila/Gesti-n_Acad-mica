from flask import request, jsonify, send_file
import sqlite3
import pandas as pd
from datetime import datetime
import os
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from .database import get_db_connection

# Configuración
UPLOAD_FOLDER = 'uploads'
REPORT_FOLDER = 'reportes'
ALLOWED_EXTENSIONS = {'csv', 'xlsx', 'xls'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def init_docente_routes(app):
    
    @app.route('/api/docentes/cargar', methods=['POST'])
    def cargar_docentes():
        try:
            if 'file' not in request.files:
                return jsonify({'success': False, 'message': 'No se encontró el archivo'}), 400
            
            file = request.files['file']
            
            if file.filename == '':
                return jsonify({'success': False, 'message': 'Nombre de archivo vacío'}), 400
            
            if not allowed_file(file.filename):
                return jsonify({'success': False, 'message': 'Formato de archivo no permitido'}), 400
            
            filename = secure_filename(file.filename)
            filepath = os.path.join(UPLOAD_FOLDER, filename)
            file.save(filepath)
            
            # Leer archivo
            if filename.endswith('.csv'):
                df = pd.read_csv(filepath, encoding='utf-8-sig')
            else:
                df = pd.read_excel(filepath)
            
            print("Columnas originales del archivo:", df.columns.tolist())
            
            # Normalizar nombres de columnas
            df.columns = [col.strip().lower().replace(' ', '_').replace('.', '').replace('(', '').replace(')', '').replace('á', 'a').replace('é', 'e').replace('í', 'i').replace('ó', 'o').replace('ú', 'u') for col in df.columns]
            
            print("Columnas normalizadas:", df.columns.tolist())
            
            # Mapeo de columnas ampliado
            mapeo_columnas = {
                'nombre_docente': ['nombre_docente', 'nombredocente', 'nombre', 'docente', 'nombre_del_docente'],
                'asignatura': ['asignatura', 'materia', 'curso', 'clase'],
                'codigo': ['codigo', 'código', 'codigo_asignatura', 'id_asignatura'],
                'dia_semana': ['dia_de_la_semana', 'dia_semana', 'diasemana', 'dia'],
                'horario_inicio': ['horario_hi', 'horario_inicio', 'hora_inicio', 'hora_entrada', 'inicio'],
                'horario_fin': ['horario_hf', 'horario_fin', 'hora_fin', 'hora_salida', 'fin'],
                'link_clase': ['link_de_la_clase', 'link_clase', 'enlace', 'url_clase', 'meet'],
                'fecha': ['fecha', 'fecha_clase', 'dia'],
                'duracion': ['duracion', 'duración', 'tiempo'],
                'link_grabacion': ['link_de_grabacion', 'link_grabacion', 'grabacion', 'url_grabacion'],
                'asistencia_estudiantes': ['asistencia_cantidad_de_estudiantes_conectados', 'asistencia_estudiantes', 'estudiantes_conectados', 'cantidad_estudiantes'],
                'asistencia_reportados': ['asistencia_cantidad_de_estudiantes_conectados_reportados', 'asistencia_reportados', 'estudiantes_reportados'],
                'asistencia_docente': ['asistencia_docente', 'docente_asistio', 'presente_docente'],
                'dependencia': ['dependencia', 'area', 'departamento', 'facultad', 'dependencia_encargada'],
                'accion': ['accion', 'proceso', 'gestion', 'actividad', 'acción'],
                'estado': ['estado', 'status', 'situacion', 'estado_proceso'],
                'evidencia': ['evidencia', 'prueba', 'soporte', 'documento', 'evidencia_proceso'],
                'observaciones': ['observaciones', 'obs', 'notas', 'comentarios']
            }
            
            renombrar = {}
            for col_destino, posibles_nombres in mapeo_columnas.items():
                for nombre in posibles_nombres:
                    if nombre in df.columns:
                        renombrar[nombre] = col_destino
                        print(f"Mapeado: '{nombre}' -> '{col_destino}'")
                        break
            
            df.rename(columns=renombrar, inplace=True)
            
            print("Columnas después del mapeo:", df.columns.tolist())
            
            # Columnas requeridas
            columnas_requeridas = ['nombre_docente', 'asignatura', 'codigo', 'dia_semana', 'horario_inicio', 'horario_fin', 'fecha']
            columnas_faltantes = [col for col in columnas_requeridas if col not in df.columns]
            if columnas_faltantes:
                return jsonify({'success': False, 'message': f'Columnas faltantes: {", ".join(columnas_faltantes)}'}), 400
            
            # Agregar columnas que no existen
            for col in ['dependencia', 'accion', 'estado', 'evidencia', 'observaciones', 'link_clase', 'duracion', 'link_grabacion', 'asistencia_estudiantes', 'asistencia_reportados', 'asistencia_docente']:
                if col not in df.columns:
                    df[col] = ''
            
            conn = get_db_connection()
            registros_insertados = 0
            
            for index, row in df.iterrows():
                cursor = conn.cursor()
                
                # Procesar valores numéricos
                asistencia_est = row.get('asistencia_estudiantes', 0)
                if pd.isna(asistencia_est):
                    asistencia_est = 0
                else:
                    try:
                        asistencia_est = int(asistencia_est)
                    except:
                        asistencia_est = 0
                
                asistencia_rep = row.get('asistencia_reportados', 0)
                if pd.isna(asistencia_rep):
                    asistencia_rep = 0
                else:
                    try:
                        asistencia_rep = int(asistencia_rep)
                    except:
                        asistencia_rep = 0
                
                cursor.execute('''
                    INSERT INTO docentes (
                        nombre_docente, asignatura, codigo, dia_semana,
                        horario_inicio, horario_fin, link_clase, fecha,
                        duracion, link_grabacion, asistencia_estudiantes,
                        asistencia_reportados, asistencia_docente, dependencia,
                        accion, estado, evidencia, observaciones
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    str(row.get('nombre_docente', '')),
                    str(row.get('asignatura', '')),
                    str(row.get('codigo', '')),
                    str(row.get('dia_semana', '')),
                    str(row.get('horario_inicio', '')),
                    str(row.get('horario_fin', '')),
                    str(row.get('link_clase', '')),
                    str(row.get('fecha', '')),
                    str(row.get('duracion', '')),
                    str(row.get('link_grabacion', '')),
                    asistencia_est,
                    asistencia_rep,
                    str(row.get('asistencia_docente', '')),
                    str(row.get('dependencia', '')),
                    str(row.get('accion', '')),
                    str(row.get('estado', 'Pendiente')),
                    str(row.get('evidencia', '')),
                    str(row.get('observaciones', ''))
                ))
                registros_insertados += 1
            
            conn.commit()
            conn.close()
            os.remove(filepath)
            
            return jsonify({
                'success': True,
                'message': f'Carga exitosa. {registros_insertados} registros insertados.'
            })
        except Exception as e:
            print(f"Error en carga masiva: {str(e)}")
            import traceback
            traceback.print_exc()
            return jsonify({'success': False, 'message': str(e)}), 500

    @app.route('/api/docentes', methods=['GET'])
    def obtener_docentes():
        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM docentes ORDER BY fecha DESC, id DESC')
            docentes = [dict(row) for row in cursor.fetchall()]
            conn.close()
            return jsonify({'success': True, 'data': docentes, 'total': len(docentes)})
        except Exception as e:
            return jsonify({'success': False, 'message': str(e)}), 500

    @app.route('/api/docentes/<int:id>', methods=['GET'])
    def obtener_docente(id):
        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM docentes WHERE id = ?', (id,))
            docente = cursor.fetchone()
            conn.close()
            
            if not docente:
                return jsonify({'success': False, 'message': 'Docente no encontrado'}), 404
            
            return jsonify({'success': True, 'data': dict(docente)})
        except Exception as e:
            return jsonify({'success': False, 'message': str(e)}), 500

    @app.route('/api/docentes/<int:id>', methods=['PUT'])
    def actualizar_docente(id):
        try:
            data = request.json
            conn = get_db_connection()
            cursor = conn.cursor()
            
            cursor.execute('''
                UPDATE docentes 
                SET nombre_docente=?, asignatura=?, codigo=?, dia_semana=?,
                    horario_inicio=?, horario_fin=?, link_clase=?, fecha=?,
                    duracion=?, link_grabacion=?, asistencia_estudiantes=?,
                    asistencia_reportados=?, asistencia_docente=?, dependencia=?,
                    accion=?, estado=?, evidencia=?, observaciones=?
                WHERE id=?
            ''', (
                data.get('nombre_docente', ''),
                data.get('asignatura', ''),
                data.get('codigo', ''),
                data.get('dia_semana', ''),
                data.get('horario_inicio', ''),
                data.get('horario_fin', ''),
                data.get('link_clase', ''),
                data.get('fecha', ''),
                data.get('duracion', ''),
                data.get('link_grabacion', ''),
                int(data.get('asistencia_estudiantes', 0)),
                int(data.get('asistencia_reportados', 0)),
                data.get('asistencia_docente', ''),
                data.get('dependencia', ''),
                data.get('accion', ''),
                data.get('estado', 'Pendiente'),
                data.get('evidencia', ''),
                data.get('observaciones', ''),
                id
            ))
            conn.commit()
            conn.close()
            return jsonify({'success': True, 'message': 'Docente actualizado correctamente'})
        except Exception as e:
            return jsonify({'success': False, 'message': str(e)}), 500

    @app.route('/api/docentes/reporte', methods=['GET'])
    def descargar_reporte_docentes():
        try:
            conn = get_db_connection()
            df = pd.read_sql_query('SELECT * FROM docentes ORDER BY fecha DESC, id DESC', conn)
            conn.close()
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Reporte Docentes"
            
            header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell_alignment = Alignment(vertical='center', wrap_text=True)
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                top=Side(style='thin'), bottom=Side(style='thin'))
            
            columns = ['ID', 'Nombre Docente', 'Asignatura', 'Código', 'Día Semana',
                       'Hora Inicio', 'Hora Fin', 'Link Clase', 'Fecha', 'Duración',
                       'Link Grabación', 'Asistencia Estudiantes', 'Asistencia Reportados',
                       'Asistencia Docente', 'Dependencia', 'Acción', 'Estado', 
                       'Evidencia', 'Observaciones', 'Fecha Creación']
            
            for col, header in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            for row_idx, row in df.iterrows():
                ws.cell(row=row_idx + 2, column=1, value=row['id'])
                ws.cell(row=row_idx + 2, column=2, value=row['nombre_docente'] or '')
                ws.cell(row=row_idx + 2, column=3, value=row['asignatura'] or '')
                ws.cell(row=row_idx + 2, column=4, value=row['codigo'] or '')
                ws.cell(row=row_idx + 2, column=5, value=row['dia_semana'] or '')
                ws.cell(row=row_idx + 2, column=6, value=row['horario_inicio'] or '')
                ws.cell(row=row_idx + 2, column=7, value=row['horario_fin'] or '')
                ws.cell(row=row_idx + 2, column=8, value=row['link_clase'] or '')
                ws.cell(row=row_idx + 2, column=9, value=row['fecha'] or '')
                ws.cell(row=row_idx + 2, column=10, value=row['duracion'] or '')
                ws.cell(row=row_idx + 2, column=11, value=row['link_grabacion'] or '')
                ws.cell(row=row_idx + 2, column=12, value=row['asistencia_estudiantes'] or 0)
                ws.cell(row=row_idx + 2, column=13, value=row['asistencia_reportados'] or 0)
                ws.cell(row=row_idx + 2, column=14, value=row['asistencia_docente'] or '')
                ws.cell(row=row_idx + 2, column=15, value=row['dependencia'] or '')
                ws.cell(row=row_idx + 2, column=16, value=row['accion'] or '')
                ws.cell(row=row_idx + 2, column=17, value=row['estado'] or 'Pendiente')
                ws.cell(row=row_idx + 2, column=18, value=row['evidencia'] or '')
                ws.cell(row=row_idx + 2, column=19, value=row['observaciones'] or '')
                ws.cell(row=row_idx + 2, column=20, value=row['fecha_creacion'] or '')
            
            # Ajustar anchos de columna
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 35)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            filename = f'reporte_docentes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            filepath = os.path.join(REPORT_FOLDER, filename)
            wb.save(filepath)
            
            return send_file(filepath, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                            as_attachment=True, download_name=filename)
        except Exception as e:
            return jsonify({'success': False, 'message': str(e)}), 500

    @app.route('/api/docentes/plantilla', methods=['GET'])
    def descargar_plantilla_docentes():
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Plantilla Docentes"
            
            header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            columns = ['Nombre Docente', 'Asignatura', 'Codigo', 'Dia Semana',
                       'Horario Inicio', 'Horario Fin', 'Link Clase', 'Fecha',
                       'Duracion', 'Link Grabacion', 'Asistencia Estudiantes',
                       'Asistencia Reportados', 'Asistencia Docente', 'Dependencia',
                       'Accion', 'Estado', 'Evidencia', 'Observaciones']
            
            for col, header in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Ejemplo de datos
            ejemplo = ['Juan Pérez', 'Matemáticas', 'MAT101', 'Lunes', '08:00', '10:00',
                      'https://meet.google.com/xxx', '2024-01-15', '2 horas',
                      'https://drive.google.com/xxx', '25', '25', 'Presente',
                      'Coordinación Académica', 'Revisión de clase', 'Completado',
                      'Acta_001.pdf', 'Sin observaciones']
            
            for col, value in enumerate(ejemplo, 1):
                ws.cell(row=2, column=col, value=value)
            
            # Ajustar anchos
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            filename = 'plantilla_docentes.xlsx'
            filepath = os.path.join(REPORT_FOLDER, filename)
            wb.save(filepath)
            
            return send_file(filepath, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                            as_attachment=True, download_name='Plantilla_Carga_Docentes.xlsx')
        except Exception as e:
            return jsonify({'success': False, 'message': str(e)}), 500

    @app.route('/api/docentes/actualizar-estado/<int:id>', methods=['PATCH'])
    def actualizar_estado_docente(id):
        try:
            data = request.json
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute('''
                UPDATE docentes 
                SET estado=?, accion=?, evidencia=?
                WHERE id=?
            ''', (
                data.get('estado', 'Pendiente'),
                data.get('accion', ''),
                data.get('evidencia', ''),
                id
            ))
            conn.commit()
            conn.close()
            return jsonify({'success': True, 'message': 'Estado actualizado correctamente'})
        except Exception as e:
            return jsonify({'success': False, 'message': str(e)}), 500

    @app.route('/api/docentes/eliminar/<int:id>', methods=['DELETE'])
    def eliminar_docente(id):
        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute('DELETE FROM docentes WHERE id = ?', (id,))
            conn.commit()
            conn.close()
            return jsonify({'success': True, 'message': 'Registro eliminado exitosamente'})
        except Exception as e:
            return jsonify({'success': False, 'message': str(e)}), 500